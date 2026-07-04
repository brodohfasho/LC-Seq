# src/core/spreadsheet_loader.py
"""
Spreadsheet loading functionality for LC-Seq.
"""

import logging
import pandas as pd
from pathlib import Path
from typing import Optional, Tuple, List

logger = logging.getLogger(__name__)


def _excel_engine(file_path: Path) -> str:
    """Return the pandas Excel engine for a workbook path."""
    return "openpyxl" if file_path.suffix.lower() == ".xlsx" else "xlrd"


class SpreadsheetLoader:
    """
    Handles loading and validation of spreadsheet files.
    
    Supports:
    - Excel files (.xlsx, .xls)
    - CSV files (.csv)
    """
    
    SUPPORTED_EXTENSIONS = {'.xlsx', '.xls', '.csv'}
    
    def __init__(self):
        """Initialize spreadsheet loader."""
        self.current_data: Optional[pd.DataFrame] = None
        self.current_file_path: Optional[Path] = None
        self.current_sheet_name: Optional[str] = None
    
    def is_supported_file(self, file_path: str) -> bool:
        """
        Check if file extension is supported.
        
        Args:
            file_path: Path to the file
            
        Returns:
            True if file extension is supported
        """
        path = Path(file_path)
        return path.suffix.lower() in self.SUPPORTED_EXTENSIONS
    
    def load_file(self, file_path: str, sheet_name: Optional[str] = None) -> Tuple[bool, Optional[str], Optional[pd.DataFrame]]:
        """
        Load a spreadsheet file.
        
        Args:
            file_path: Path to the spreadsheet file
            sheet_name: Name of Excel sheet to load (None for first sheet or CSV)
            
        Returns:
            Tuple of (success, error_message, dataframe)
            - success: True if load was successful
            - error_message: Error message if failed, None if successful
            - dataframe: Loaded dataframe if successful, None if failed
        """
        file_path_obj = Path(file_path)
        
        # Validate file exists
        if not file_path_obj.exists():
            error = f"File not found: {file_path}"
            logger.error(error)
            return False, error, None
        
        # Validate file extension
        if not self.is_supported_file(file_path):
            error = f"Unsupported file type: {file_path_obj.suffix}. Supported: {', '.join(self.SUPPORTED_EXTENSIONS)}"
            logger.error(error)
            return False, error, None
        
        try:
            # Load based on file type
            if file_path_obj.suffix.lower() == '.csv':
                df = pd.read_csv(file_path, encoding='utf-8')
                self.current_sheet_name = None
            elif file_path_obj.suffix.lower() in ['.xlsx', '.xls']:
                # Handle Excel files
                engine = _excel_engine(file_path_obj)
                if sheet_name:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, engine=engine)
                else:
                    # Load first sheet by default
                    excel_file = pd.ExcelFile(file_path, engine=engine)
                    sheet_name = excel_file.sheet_names[0]
                    df = pd.read_excel(file_path, sheet_name=sheet_name, engine=engine)
                self.current_sheet_name = sheet_name
            else:
                error = f"Unexpected file type: {file_path_obj.suffix}"
                logger.error(error)
                return False, error, None
            
            # Validate basic structure
            if df.empty:
                error = "Spreadsheet is empty (no data rows)"
                logger.error(error)
                return False, error, None
            
            if df.shape[1] == 0:
                error = "Spreadsheet has no columns"
                logger.error(error)
                return False, error, None
            
            # Store loaded data
            self.current_data = df
            self.current_file_path = file_path_obj
            
            logger.info(f"Successfully loaded spreadsheet: {file_path} ({df.shape[0]} rows, {df.shape[1]} columns)")
            return True, None, df
        
        except pd.errors.EmptyDataError:
            error = "Spreadsheet appears to be empty or corrupted"
            logger.error(f"{error}: {file_path}")
            return False, error, None
        
        except Exception as e:
            error = f"Error loading spreadsheet: {str(e)}"
            logger.error(f"{error}: {file_path}", exc_info=True)
            return False, error, None
    
    def get_available_sheets(self, file_path: str) -> Optional[List[str]]:
        """
        Get list of available sheet names for an Excel file.

        Uses a read-only openpyxl pass for ``.xlsx`` files so sheet detection
        does not parse the full workbook into memory.

        Args:
            file_path: Path to Excel file

        Returns:
            List of sheet names, or None if not an Excel file or error occurred
        """
        file_path_obj = Path(file_path)

        if file_path_obj.suffix.lower() not in ['.xlsx', '.xls']:
            return None

        try:
            if file_path_obj.suffix.lower() == '.xlsx':
                from openpyxl import load_workbook

                workbook = load_workbook(file_path, read_only=True, data_only=True)
                try:
                    return list(workbook.sheetnames)
                finally:
                    workbook.close()
            excel_file = pd.ExcelFile(file_path, engine='xlrd')
            return excel_file.sheet_names
        except Exception as e:
            logger.error(f"Error reading Excel sheets: {e}")
            return None
    
    def get_column_names(self) -> List[str]:
        """
        Get column names from currently loaded spreadsheet.
        
        Returns:
            List of column names, empty list if no data loaded
        """
        if self.current_data is None:
            return []
        return list(self.current_data.columns)
    
    def validate_required_columns(self, required_columns: List[str]) -> Tuple[bool, Optional[str]]:
        """
        Validate that required columns exist in loaded spreadsheet.
        
        Args:
            required_columns: List of required column names
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        if self.current_data is None:
            return False, "No spreadsheet loaded"
        
        available_columns = self.get_column_names()
        missing_columns = [col for col in required_columns if col not in available_columns]
        
        if missing_columns:
            error = f"Missing required columns: {', '.join(missing_columns)}"
            logger.warning(error)
            return False, error
        
        return True, None
    
    def get_data(self) -> Optional[pd.DataFrame]:
        """
        Get the currently loaded dataframe.
        
        Returns:
            Current dataframe, or None if no data loaded
        """
        return self.current_data
    
    def clear_data(self) -> None:
        """Clear currently loaded data."""
        self.current_data = None
        self.current_file_path = None
        self.current_sheet_name = None
        logger.debug("Cleared loaded spreadsheet data")
