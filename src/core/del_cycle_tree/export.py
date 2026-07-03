# src/core/del_cycle_tree/export.py
"""CSV and Excel export for DEL-cycle split-tree analysis."""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional, Sequence, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.core.del_cycle_tree.bb_index_scheme import lookup_bb_display_index
from src.core.del_cycle_tree.models import DelCycleTreeData, VerifiedSequence
from src.core.pedigree_export import bb_cycle_field_names

_PASS_FILL = PatternFill(start_color="00CC00", end_color="00CC00", fill_type="solid")
_FAIL_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
_MAJOR_FAIL_PASS_PCT = 50.0

ProgressCallback = Callable[[float, str], None]

_AUDIT_FIELDS = (
    "rt_threshold",
    "rt_source",
    "peak_picking_algorithm",
    "n_rt_from_pedigree",
    "n_rt_from_peak_pick",
    "n_rt_from_metadata",
    "n_rt_verified_pedigree_agree",
    "full_null_rt",
    "n_products",
    "n_rt_verified",
    "n_pedigree_passed",
)


@dataclass(frozen=True)
class DelCycleExportResult:
    """Paths written by :func:`export_del_cycle_package`."""

    output_dir: Path
    products_csv: Path
    audit_csv: Path
    summary_csv: Path
    flagged_csv: Path
    grid_files: Tuple[Path, ...] = field(default_factory=tuple)

    @property
    def file_count(self) -> int:
        return 4 + len(self.grid_files)


def _report_progress(
    callback: Optional[ProgressCallback],
    fraction: float,
    status: str,
) -> None:
    if callback is not None:
        callback(min(1.0, max(0.0, fraction)), status)


def _bool_csv(value: bool) -> str:
    return "TRUE" if value else "FALSE"


def _safe_filename(name: str) -> str:
    text = re.sub(r"[^\w\-]+", "_", name, flags=re.UNICODE).strip("_")
    return text or "unnamed"


def _products_fieldnames(library_cycle_count: int) -> List[str]:
    cycles = bb_cycle_field_names(library_cycle_count)
    index_cols = [f"bb{k}_index" for k in range(1, 5)]
    return [
        *cycles,
        "rt (s)",
        "rt_verified",
        "pedigree_passed",
        *index_cols,
    ]


def _cycle_indices(
    positions: Tuple[str, ...],
    index_map: Dict[str, int],
    *,
    null_token: str,
) -> Dict[str, int | None]:
    out: Dict[str, int | None] = {}
    for cycle in range(1, 5):
        key = f"bb{cycle}_index"
        if cycle - 1 < len(positions):
            out[key] = lookup_bb_display_index(
                positions[cycle - 1],
                index_map,
                null_token=null_token,
            )
        else:
            out[key] = None
    return out


def _cycle_columns(positions: Tuple[str, ...], n_cycles: int) -> Dict[str, str]:
    cols = bb_cycle_field_names(n_cycles)
    out = {name: "" for name in cols}
    for index, name in enumerate(cols):
        if index < len(positions):
            out[name] = positions[index]
    return out


def _audit_metadata(data: DelCycleTreeData, *, n_products: int) -> Dict[str, object]:
    return {
        "rt_threshold": data.rt_threshold,
        "rt_source": data.rt_source,
        "peak_picking_algorithm": data.peak_picking_algorithm,
        "n_rt_from_pedigree": data.n_rt_from_pedigree,
        "n_rt_from_peak_pick": data.n_rt_from_peak_pick,
        "n_rt_from_metadata": data.n_rt_from_metadata,
        "n_rt_verified_pedigree_agree": data.n_rt_verified_pedigree_agree,
        "full_null_rt": data.full_null_rt if data.full_null_rt is not None else "",
        "n_products": n_products,
        "n_rt_verified": data.n_verified,
        "n_pedigree_passed": data.n_pedigree_passed,
    }


def _iter_full_products(
    data: DelCycleTreeData,
) -> Iterable[Tuple[Tuple[str, ...], VerifiedSequence]]:
    n_cycles = data.library_cycle_count
    null = data.null_token
    for positions, info in sorted(data.verified_sequences.items(), key=lambda item: item[0]):
        if len(positions) != n_cycles:
            continue
        if any(bb == null for bb in positions):
            continue
        yield positions, info


def _product_row(data: DelCycleTreeData, positions: Tuple[str, ...], info: VerifiedSequence) -> Dict[str, object]:
    indices = _cycle_indices(positions, data.bb_index_global, null_token=data.null_token)
    pedigree_passed = data.pedigree_passed_by_product.get(positions)
    row: Dict[str, object] = {
        "rt (s)": info.rt,
        "rt_verified": _bool_csv(bool(info.success)),
        "pedigree_passed": (
            _bool_csv(bool(pedigree_passed))
            if pedigree_passed is not None
            else ""
        ),
        "bb1_index": indices["bb1_index"] if indices["bb1_index"] is not None else "",
        "bb2_index": indices["bb2_index"] if indices["bb2_index"] is not None else "",
        "bb3_index": indices["bb3_index"] if indices["bb3_index"] is not None else "",
        "bb4_index": indices["bb4_index"] if indices["bb4_index"] is not None else "",
    }
    row.update(_cycle_columns(positions, data.library_cycle_count))
    return row


def _unique_bbs_at_depth(data: DelCycleTreeData, depth: int) -> List[str]:
    """Collect unique non-null BB names at coupling depth ``depth`` (0 = BB1)."""
    null = data.null_token
    names: Set[str] = set()

    def walk(node: object, current_depth: int) -> None:
        if not isinstance(node, dict):
            return
        for name, child in node.items():
            if name == null:
                continue
            if current_depth == depth:
                names.add(name)
            elif isinstance(child, dict):
                walk(child, current_depth + 1)

    walk(data.tree, 0)
    for positions, _info in _iter_full_products(data):
        if len(positions) > depth and positions[depth] != null:
            names.add(positions[depth])
    return sorted(names, key=lambda n: (data.bb_index_global.get(n, 10**9), n.lower()))


def _bb1_names_for_export(data: DelCycleTreeData) -> List[str]:
    null = data.null_token
    return [name for name in data.bb1_names if name != null]


def _write_products_csv(data: DelCycleTreeData, path: Path) -> None:
    fieldnames = _products_fieldnames(data.library_cycle_count)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for positions, info in _iter_full_products(data):
            writer.writerow(_product_row(data, positions, info))


def _write_audit_csv(data: DelCycleTreeData, path: Path, *, n_products: int) -> None:
    meta = _audit_metadata(data, n_products=n_products)
    exported_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["field", "value"])
        writer.writerow(["exported_at_utc", exported_at])
        writer.writerow(["library_cycle_count", data.library_cycle_count])
        writer.writerow(["null_token", data.null_token])
        for key in _AUDIT_FIELDS:
            writer.writerow([key, meta[key]])


def _prefix_stats(
    data: DelCycleTreeData,
    prefix: Tuple[str, ...],
) -> Tuple[int, int, int]:
    """Return (total, n_pass, n_fail) for full products under ``prefix``."""
    total = 0
    passed = 0
    for positions, info in _iter_full_products(data):
        if positions[: len(prefix)] != prefix:
            continue
        total += 1
        if info.success:
            passed += 1
    return total, passed, total - passed


def _build_summary_rows(data: DelCycleTreeData) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    null = data.null_token

    for bb1 in _bb1_names_for_export(data):
        total, passed, failed = _prefix_stats(data, (bb1,))
        pass_pct = (100.0 * passed / total) if total else 0.0
        rows.append(
            {
                "scope": "cycle_1",
                "bb_cycle_1": bb1,
                "bb_cycle_2": "",
                "bb1_index": lookup_bb_display_index(bb1, data.bb_index_global, null_token=null) or "",
                "bb2_index": "",
                "total_products": total,
                "n_rt_verified_pass": passed,
                "n_rt_verified_fail": failed,
                "pass_pct": round(pass_pct, 1),
                "majority_failed": _bool_csv(total > 0 and pass_pct < _MAJOR_FAIL_PASS_PCT),
                "flag_reason": (
                    "Majority of descendant products failed RT verification"
                    if total > 0 and pass_pct < _MAJOR_FAIL_PASS_PCT
                    else ""
                ),
            }
        )

    bb2_names = _unique_bbs_at_depth(data, 1)
    for bb2 in bb2_names:
        if bb2 == null:
            continue
        total = 0
        passed = 0
        for positions, info in _iter_full_products(data):
            if len(positions) < 2 or positions[1] != bb2:
                continue
            total += 1
            if info.success:
                passed += 1
        failed = total - passed
        pass_pct = (100.0 * passed / total) if total else 0.0
        rows.append(
            {
                "scope": "cycle_2",
                "bb_cycle_1": "",
                "bb_cycle_2": bb2,
                "bb1_index": "",
                "bb2_index": lookup_bb_display_index(bb2, data.bb_index_global, null_token=null) or "",
                "total_products": total,
                "n_rt_verified_pass": passed,
                "n_rt_verified_fail": failed,
                "pass_pct": round(pass_pct, 1),
                "majority_failed": _bool_csv(total > 0 and pass_pct < _MAJOR_FAIL_PASS_PCT),
                "flag_reason": (
                    "Majority of products containing this cycle-2 BB failed RT verification"
                    if total > 0 and pass_pct < _MAJOR_FAIL_PASS_PCT
                    else ""
                ),
            }
        )

    for bb1 in _bb1_names_for_export(data):
        bb2_under = _unique_bbs_at_depth(data, 1)
        for bb2 in bb2_under:
            if bb2 == null:
                continue
            total, passed, failed = _prefix_stats(data, (bb1, bb2))
            if total == 0:
                continue
            pass_pct = 100.0 * passed / total
            if pass_pct >= _MAJOR_FAIL_PASS_PCT:
                continue
            rows.append(
                {
                    "scope": "cycle_1_and_2",
                    "bb_cycle_1": bb1,
                    "bb_cycle_2": bb2,
                    "bb1_index": lookup_bb_display_index(bb1, data.bb_index_global, null_token=null) or "",
                    "bb2_index": lookup_bb_display_index(bb2, data.bb_index_global, null_token=null) or "",
                    "total_products": total,
                    "n_rt_verified_pass": passed,
                    "n_rt_verified_fail": failed,
                    "pass_pct": round(pass_pct, 1),
                    "majority_failed": "TRUE",
                    "flag_reason": (
                        f"Under BB1={bb1}, cycle-2 arm {bb2} yields majority RT verification failures"
                    ),
                }
            )

    return rows


def _library_pass_stats(data: DelCycleTreeData) -> Tuple[int, int, float]:
    """Return (total products, n pass, pass %)."""
    total = 0
    passed = 0
    for _positions, info in _iter_full_products(data):
        total += 1
        if info.success:
            passed += 1
    pass_pct = (100.0 * passed / total) if total else 0.0
    return total, passed, pass_pct


def _format_coupling_detail(bb1: str, bb2: str, pass_pct: object, total: object) -> str:
    return f"BB1 {bb1} + BB2 {bb2} ({pass_pct}% pass, n={total})"


def _build_flagged_bb_commentary(
    *,
    coupling_cycle: int,
    bb_name: str,
    pass_pct: float,
    total_products: int,
    library_pass_pct: float,
    cycle_1_hub: bool,
    cycle_2_global: bool,
    coupling_details: Sequence[str],
    n_independent_flags: int,
) -> str:
    parts: List[str] = []
    if n_independent_flags > 1:
        parts.append(
            f"Flagged {n_independent_flags} times across independent coupling analyses."
        )
    if cycle_1_hub:
        parts.append(
            f"Majority RT verification failure as cycle-1 hub "
            f"({pass_pct:.1f}% pass across {total_products} descendant product(s))."
        )
    if cycle_2_global:
        parts.append(
            f"Majority RT verification failure when present at cycle 2 "
            f"({pass_pct:.1f}% pass across {total_products} product(s))."
        )
    if coupling_details:
        if coupling_cycle == 1:
            parts.append(
                "Problematic cycle-2 arms under this BB1: "
                + "; ".join(coupling_details)
                + "."
            )
        else:
            parts.append(
                "Problematic cycle-1 partners for this BB2: "
                + "; ".join(coupling_details)
                + "."
            )
    delta = pass_pct - library_pass_pct
    if total_products > 0:
        if delta < -10:
            parts.append(
                f"Overall pass rate is {abs(delta):.1f} percentage points below "
                f"the library average ({library_pass_pct:.1f}%)."
            )
        elif delta > 10:
            parts.append(
                f"Overall pass rate is {delta:.1f} percentage points above "
                f"the library average ({library_pass_pct:.1f}%)."
            )
    if not parts:
        parts.append(
            f"{bb_name} appeared in at least one majority-failure coupling context."
        )
    return " ".join(parts)


def _build_flagged_bb_rows(data: DelCycleTreeData) -> List[Dict[str, object]]:
    """Aggregate repeatedly flagged building blocks with commentary."""
    summary = _build_summary_rows(data)
    null = data.null_token
    _total_lib, _passed_lib, library_pass_pct = _library_pass_stats(data)

    baseline: Dict[Tuple[int, str], Dict[str, object]] = {}
    for row in summary:
        scope = row["scope"]
        if scope == "cycle_1":
            baseline[(1, str(row["bb_cycle_1"]))] = row
        elif scope == "cycle_2":
            baseline[(2, str(row["bb_cycle_2"]))] = row

    flagged = [row for row in summary if row.get("majority_failed") == "TRUE"]

    cycle_1_hub: Set[Tuple[int, str]] = set()
    cycle_2_global: Set[Tuple[int, str]] = set()
    couplings_by_bb1: Dict[str, List[str]] = {}
    couplings_by_bb2: Dict[str, List[str]] = {}

    for row in flagged:
        scope = row["scope"]
        if scope == "cycle_1":
            name = str(row["bb_cycle_1"])
            cycle_1_hub.add((1, name))
        elif scope == "cycle_2":
            name = str(row["bb_cycle_2"])
            cycle_2_global.add((2, name))
        elif scope == "cycle_1_and_2":
            bb1 = str(row["bb_cycle_1"])
            bb2 = str(row["bb_cycle_2"])
            detail = _format_coupling_detail(
                bb1,
                bb2,
                row["pass_pct"],
                row["total_products"],
            )
            couplings_by_bb1.setdefault(bb1, []).append(detail)
            partner_detail = _format_coupling_detail(
                bb1,
                bb2,
                row["pass_pct"],
                row["total_products"],
            )
            couplings_by_bb2.setdefault(bb2, []).append(partner_detail)

    flagged_keys: Set[Tuple[int, str]] = set()
    flagged_keys.update(cycle_1_hub)
    flagged_keys.update(cycle_2_global)
    flagged_keys.update((1, name) for name in couplings_by_bb1)
    flagged_keys.update((2, name) for name in couplings_by_bb2)

    rows_out: List[Dict[str, object]] = []
    for cycle, name in flagged_keys:
        if not name or name == null:
            continue
        base = baseline.get((cycle, name), {})
        total_products = int(base.get("total_products") or 0)
        n_pass = int(base.get("n_rt_verified_pass") or 0)
        n_fail = int(base.get("n_rt_verified_fail") or 0)
        pass_pct = float(base.get("pass_pct") or 0.0)
        hub = (cycle, name) in cycle_1_hub
        global_bb2 = (cycle, name) in cycle_2_global
        if cycle == 1:
            coupling_details = couplings_by_bb1.get(name, [])
        else:
            coupling_details = couplings_by_bb2.get(name, [])
        n_coupling_flags = len(coupling_details)
        n_independent_flags = int(hub) + int(global_bb2) + n_coupling_flags
        bb_index = lookup_bb_display_index(name, data.bb_index_global, null_token=null)
        rows_out.append(
            {
                "bb_name": name,
                "bb_index": bb_index if bb_index is not None else "",
                "coupling_cycle": cycle,
                "total_products": total_products,
                "n_rt_verified_pass": n_pass,
                "n_rt_verified_fail": n_fail,
                "pass_pct": round(pass_pct, 1),
                "library_pass_pct": round(library_pass_pct, 1),
                "pass_pct_vs_library": round(pass_pct - library_pass_pct, 1),
                "n_independent_flags": n_independent_flags,
                "flagged_as_cycle_1_hub": _bool_csv(hub),
                "flagged_as_cycle_2_global": _bool_csv(global_bb2),
                "n_flagged_cycle1_cycle2_couplings": n_coupling_flags,
                "flagged_coupling_details": "; ".join(coupling_details),
                "commentary": _build_flagged_bb_commentary(
                    coupling_cycle=cycle,
                    bb_name=name,
                    pass_pct=pass_pct,
                    total_products=total_products,
                    library_pass_pct=library_pass_pct,
                    cycle_1_hub=hub,
                    cycle_2_global=global_bb2,
                    coupling_details=coupling_details,
                    n_independent_flags=n_independent_flags,
                ),
            }
        )

    rows_out.sort(
        key=lambda row: (
            -int(row["n_independent_flags"]),
            float(row["pass_pct"]),
            int(row["coupling_cycle"]),
            str(row["bb_name"]).lower(),
        ),
    )
    return rows_out


def _write_flagged_bb_csv(data: DelCycleTreeData, path: Path) -> int:
    fieldnames = [
        "bb_name",
        "bb_index",
        "coupling_cycle",
        "total_products",
        "n_rt_verified_pass",
        "n_rt_verified_fail",
        "pass_pct",
        "library_pass_pct",
        "pass_pct_vs_library",
        "n_independent_flags",
        "flagged_as_cycle_1_hub",
        "flagged_as_cycle_2_global",
        "n_flagged_cycle1_cycle2_couplings",
        "flagged_coupling_details",
        "commentary",
    ]
    rows = _build_flagged_bb_rows(data)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    return len(rows)


def _write_summary_csv(data: DelCycleTreeData, path: Path) -> None:
    fieldnames = [
        "scope",
        "bb_cycle_1",
        "bb_cycle_2",
        "bb1_index",
        "bb2_index",
        "total_products",
        "n_rt_verified_pass",
        "n_rt_verified_fail",
        "pass_pct",
        "majority_failed",
        "flag_reason",
    ]
    rows = _build_summary_rows(data)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _write_bb1_grid_xlsx(
    data: DelCycleTreeData,
    *,
    bb1: str,
    bb2_names: Sequence[str],
    bb3_names: Sequence[str],
    path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RT verified"
    ws.cell(row=1, column=1, value="BB2 \\ BB3")
    header_font = Font(bold=True)
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    for col, bb3 in enumerate(bb3_names, start=2):
        cell = ws.cell(row=1, column=col, value=bb3)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, bb2 in enumerate(bb2_names, start=2):
        label = ws.cell(row=row_idx, column=1, value=bb2)
        label.font = header_font
        for col_idx, bb3 in enumerate(bb3_names, start=2):
            positions = (bb1, bb2, bb3)
            info = data.verified_sequences.get(positions)
            cell = ws.cell(row=row_idx, column=col_idx)
            if info is None:
                cell.value = ""
                continue
            cell.value = "PASS" if info.success else "FAIL"
            cell.fill = _PASS_FILL if info.success else _FAIL_FILL
            cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "B2"
    wb.save(path)


def _write_bb1_grids(
    data: DelCycleTreeData,
    grids_dir: Path,
    *,
    progress_callback: Optional[ProgressCallback] = None,
    progress_start: float = 0.30,
    progress_end: float = 0.95,
) -> List[Path]:
    if data.library_cycle_count != 3:
        return []

    bb2_names = _unique_bbs_at_depth(data, 1)
    bb3_names = _unique_bbs_at_depth(data, 2)
    if not bb2_names or not bb3_names:
        return []

    grids_dir.mkdir(parents=True, exist_ok=True)
    written: List[Path] = []
    bb1_list = _bb1_names_for_export(data)
    total = len(bb1_list)
    span = max(progress_end - progress_start, 0.0)
    for index, bb1 in enumerate(bb1_list):
        if total > 0:
            fraction = progress_start + span * (index / total)
            _report_progress(
                progress_callback,
                fraction,
                f"Writing BB1 grid {index + 1}/{total}: {bb1}…",
            )
        safe = _safe_filename(bb1)
        bb_index = lookup_bb_display_index(bb1, data.bb_index_global, null_token=data.null_token)
        prefix = f"del_grid_bb1_{bb_index}_{safe}" if bb_index else f"del_grid_bb1_{safe}"
        path = grids_dir / f"{prefix}.xlsx"
        _write_bb1_grid_xlsx(
            data,
            bb1=bb1,
            bb2_names=bb2_names,
            bb3_names=bb3_names,
            path=path,
        )
        written.append(path)
    if total > 0:
        _report_progress(
            progress_callback,
            progress_end,
            f"Finished {total} BB1 grid workbook(s).",
        )
    return written


def export_del_cycle_package(
    data: DelCycleTreeData,
    output_dir: str | Path,
    *,
    progress_callback: Optional[ProgressCallback] = None,
) -> DelCycleExportResult:
    """
    Write DEL-cycle export bundle into ``output_dir``.

    Creates:
    - ``del_cycle_products.csv`` — product table (no audit columns)
    - ``del_cycle_audit_metadata.csv`` — run metadata and audit counters
    - ``del_cycle_summary_report.csv`` — flagged BB1/BB2 majority-failure patterns
    - ``del_cycle_flagged_building_blocks.csv`` — aggregated problematic residues with commentary
    - ``grids/del_grid_bb1_*.xlsx`` — one color-coded BB2×BB3 grid per BB1 (3-cycle only)
    """
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    _report_progress(progress_callback, 0.0, "Preparing DEL-cycle export…")
    products = list(_iter_full_products(data))
    products_csv = out_dir / "del_cycle_products.csv"
    audit_csv = out_dir / "del_cycle_audit_metadata.csv"
    summary_csv = out_dir / "del_cycle_summary_report.csv"
    flagged_csv = out_dir / "del_cycle_flagged_building_blocks.csv"
    grids_dir = out_dir / "grids"

    _report_progress(progress_callback, 0.05, "Writing product table…")
    _write_products_csv(data, products_csv)
    _report_progress(progress_callback, 0.15, "Writing audit metadata…")
    _write_audit_csv(data, audit_csv, n_products=len(products))
    _report_progress(progress_callback, 0.22, "Writing summary report…")
    _write_summary_csv(data, summary_csv)
    _report_progress(progress_callback, 0.28, "Writing flagged building blocks…")
    _write_flagged_bb_csv(data, flagged_csv)
    grid_files = _write_bb1_grids(
        data,
        grids_dir,
        progress_callback=progress_callback,
        progress_start=0.32,
        progress_end=0.95,
    )
    _report_progress(progress_callback, 1.0, "Export complete.")

    return DelCycleExportResult(
        output_dir=out_dir,
        products_csv=products_csv,
        audit_csv=audit_csv,
        summary_csv=summary_csv,
        flagged_csv=flagged_csv,
        grid_files=tuple(grid_files),
    )


def export_del_cycle_csv(data: DelCycleTreeData, path: str | Path) -> Path:
    """Legacy single-file export; writes only the product table to ``path``."""
    out = Path(path)
    _write_products_csv(data, out)
    return out
