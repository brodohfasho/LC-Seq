# tests/test_del_cycle_export.py
"""Tests for DEL-cycle CSV / Excel export bundle."""

from __future__ import annotations

import csv
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.core.del_cycle_tree.export import export_del_cycle_package
from src.core.del_cycle_tree.models import DelCycleTreeData, VerifiedSequence
from src.models.analysis_settings import AnalysisSettings


def _sample_data() -> DelCycleTreeData:
    null = "AgxNull"
    return DelCycleTreeData(
        library_cycle_count=3,
        null_token=null,
        rt_threshold=0.5,
        tree={
            null: {},
            "LA03": {"DPhe": {"X": 1.0}, "Leu": {"Y": 2.0}},
            "LA18": {"DPhe": {"Z": 3.0}, "Leu": {"W": 4.0}},
        },
        pruned_tree={},
        verified_sequences={
            ("LA03", "DPhe", "X"): VerifiedSequence(("LA03", "DPhe", "X"), 1.0, True),
            ("LA03", "DPhe", "Y"): VerifiedSequence(("LA03", "DPhe", "Y"), 1.1, False),
            ("LA03", "Leu", "X"): VerifiedSequence(("LA03", "Leu", "X"), 1.2, False),
            ("LA03", "Leu", "Y"): VerifiedSequence(("LA03", "Leu", "Y"), 1.3, False),
            ("LA18", "DPhe", "Z"): VerifiedSequence(("LA18", "DPhe", "Z"), 2.0, True),
            ("LA18", "Leu", "W"): VerifiedSequence(("LA18", "Leu", "W"), 2.1, True),
        },
        full_null_rt=10.0,
        bb_index_global={"LA03": 30, "LA18": 31, "DPhe": 4, "Leu": 12, "X": 20, "Y": 21, "Z": 22, "W": 23},
        bb1_names=[null, "LA03", "LA18"],
        n_verified=3,
        rt_source="pedigree",
        peak_picking_algorithm="gaussian",
        n_rt_from_pedigree=6,
        n_rt_from_peak_pick=0,
        n_rt_from_metadata=0,
        n_rt_verified_pedigree_agree=5,
        pedigree_passed_by_product={
            ("LA03", "DPhe", "X"): True,
            ("LA03", "DPhe", "Y"): False,
            ("LA18", "DPhe", "Z"): True,
            ("LA18", "Leu", "W"): True,
        },
        n_pedigree_passed=3,
    )


def test_export_del_cycle_package_writes_all_artifacts(tmp_path: Path) -> None:
    data = _sample_data()
    result = export_del_cycle_package(data, tmp_path)

    assert result.products_csv.is_file()
    assert result.audit_csv.is_file()
    assert result.summary_csv.is_file()
    assert result.flagged_csv.is_file()
    assert len(result.grid_files) == 2
    assert result.file_count == 6


def test_products_csv_schema(tmp_path: Path) -> None:
    data = _sample_data()
    result = export_del_cycle_package(data, tmp_path)

    with result.products_csv.open(encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh))

    assert len(rows) == 6
    assert "row_kind" not in rows[0]
    assert "rt (s)" in rows[0]
    assert "rt_threshold" not in rows[0]
    assert rows[0]["rt_verified"] in {"TRUE", "FALSE"}
    assert rows[0]["pedigree_passed"] in {"TRUE", "FALSE", ""}


def test_audit_metadata_csv(tmp_path: Path) -> None:
    data = _sample_data()
    settings = AnalysisSettings(
        count_channel="signal",
        time_unit="seconds",
        peak_picking_algorithm="old_school",
        tolerance=30.0,
        gaussian_minimum_rt=600.0,
        gaussian_fit_width=90.0,
        gaussian_stddev_threshold=120.0,
    )
    result = export_del_cycle_package(
        data,
        tmp_path,
        analysis_settings=settings,
        rt_analysis_mode="direct_pick",
    )

    with result.audit_csv.open(encoding="utf-8", newline="") as fh:
        audit = {row["field"]: row["value"] for row in csv.DictReader(fh)}

    assert audit["analysis_time_unit"] == "seconds"
    assert audit["null_rt_threshold"] == "30.0"
    assert audit["null_rt_threshold_unit"] == "seconds"
    assert audit["gaussian_minimum_rt_seconds"] == "600.0"
    assert audit["rt_threshold"] == "0.5"
    assert audit["rt_source"] == "pedigree"
    assert audit["rt_analysis_mode"] == "direct_pick"
    assert audit["n_rt_from_pedigree"] == "6"
    assert audit["n_products"] == "6"
    assert audit["full_null_rt"] == "10.0"

    with result.products_csv.open(encoding="utf-8", newline="") as fh:
        products = list(csv.DictReader(fh))
    assert "rt (s)" in products[0]


def test_products_csv_rt_column_minutes(tmp_path: Path) -> None:
    data = _sample_data()
    settings = AnalysisSettings(
        count_channel="signal",
        time_unit="minutes",
        peak_picking_algorithm="modern",
        tolerance=0.5,
    )
    result = export_del_cycle_package(
        data,
        tmp_path,
        analysis_settings=settings,
        rt_analysis_mode="direct_pick",
    )

    with result.products_csv.open(encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh))

    assert rows
    assert "rt (min)" in rows[0]
    assert "rt (s)" not in rows[0]


def test_summary_flags_majority_failed_bb1(tmp_path: Path) -> None:
    data = _sample_data()
    result = export_del_cycle_package(data, tmp_path)

    with result.summary_csv.open(encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh))

    la03 = next(r for r in rows if r["scope"] == "cycle_1" and r["bb_cycle_1"] == "LA03")
    assert la03["majority_failed"] == "TRUE"
    assert float(la03["pass_pct"]) < 50.0

    la18 = next(r for r in rows if r["scope"] == "cycle_1" and r["bb_cycle_1"] == "LA18")
    assert la18["majority_failed"] == "FALSE"


def test_bb1_grid_workbook_colors(tmp_path: Path) -> None:
    data = _sample_data()
    result = export_del_cycle_package(data, tmp_path)
    la03_grid = next(p for p in result.grid_files if "LA03" in p.name)

    wb = load_workbook(la03_grid)
    ws = wb.active
    assert ws.cell(row=2, column=2).value == "PASS"
    assert ws.cell(row=2, column=2).fill.start_color.rgb in {"0000CC00", "00CC00"}
    assert ws.cell(row=2, column=3).value == "FAIL"
    assert ws.cell(row=2, column=3).fill.start_color.rgb in {"00FF0000", "FF0000"}


def test_export_reports_progress(tmp_path: Path) -> None:
    data = _sample_data()
    updates: list[tuple[float, str]] = []

    def on_progress(fraction: float, status: str) -> None:
        updates.append((fraction, status))

    export_del_cycle_package(data, tmp_path, progress_callback=on_progress)
    assert updates
    assert updates[0][0] == 0.0
    assert updates[-1][0] == 1.0
    assert any("grid" in status.lower() for _fraction, status in updates)


def test_no_grids_for_two_cycle_library(tmp_path: Path) -> None:
    data = _sample_data()
    two_cycle = DelCycleTreeData(
        library_cycle_count=2,
        null_token=data.null_token,
        rt_threshold=data.rt_threshold,
        tree={"LA03": {"DPhe": 1.0}},
        pruned_tree={},
        verified_sequences={
            ("LA03", "DPhe"): VerifiedSequence(("LA03", "DPhe"), 1.0, True),
        },
        full_null_rt=10.0,
        bb_index_global={"LA03": 30, "DPhe": 4},
        bb1_names=["LA03"],
    )
    result = export_del_cycle_package(two_cycle, tmp_path)
    assert result.grid_files == ()
    assert result.file_count == 4


def test_del_cycle_bundle_glossary_help_topic() -> None:
    from src.core.help_content import load_help_text

    text = load_help_text("del_cycle_bundle_glossary")
    assert "del_cycle_products.csv" in text
    assert "rt_verified" in text
    assert "grids/" in text
    assert "green fill" in text
    assert "del_cycle_flagged_building_blocks.csv" in text
    assert "product_prominence.csv" in text
    assert "compound_id" in text


def test_flagged_building_blocks_csv(tmp_path: Path) -> None:
    data = _sample_data()
    result = export_del_cycle_package(data, tmp_path)

    with result.flagged_csv.open(encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh))

    assert rows
    la03 = next(r for r in rows if r["bb_name"] == "LA03" and r["coupling_cycle"] == "1")
    assert la03["flagged_as_cycle_1_hub"] == "TRUE"
    assert int(la03["n_independent_flags"]) >= 2
    assert "Flagged" in la03["commentary"]
    assert float(la03["pass_pct"]) < 50.0
    assert "library average" in la03["commentary"].lower()

    leu = next(r for r in rows if r["bb_name"] == "Leu" and r["coupling_cycle"] == "2")
    assert int(leu["n_independent_flags"]) >= 1
    assert leu["flagged_coupling_details"]
